from pathlib import Path
import json
from db import CteCRUD
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

nomes_campos_colunas = {
    "dataEmissao": "Emissão",
    "tipoDocumento": "Tipo do documento",
    "idDocumentoFiscal": "Chave do CT-e",
    "numeroDocumento": "Número do CT-e",
    "numeroDocumentoReferenciado": "Número do CT-e referenciado",
    "nr_cnpj_cpf_pagador": "Adquirinte",
    "cd_mun_orig": "Cidade origem",
    "uf_orig": "UF Origem",
    "cd_mun_dest": "Cidade Destino",
    "uf_dest": "UF Destino",
    "vl_total": "Valor da operação",
    # "seguro": "Seguro",
    # "frete": "Frete",
    "valorLancamento": "Valor original",
    "valorCalculado": "Valor calculado RFB",
    "ValorNaoPago": "Saldo a pagar"
}


def extrair_id_documentos_fiscais(arquivo_json):
    """
    Lê o arquivo JSON e extrai todos os valores do campo idDocumentoFiscal.
    
    Args:
        arquivo_json: Caminho para o arquivo JSON
        
    Returns:
        Tupla com todos os valores de idDocumentoFiscal
    """
    caminho = Path(arquivo_json)
    
    with open(caminho, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    
    id_documentos = tuple(item['idDocumentoFiscal'] for item in dados['dados'])
    
    return id_documentos


def carregar_municipios(arquivo_municipios='municipios.json'):
    """
    Carrega o arquivo de municípios e cria um dicionário indexado por código.
    
    Args:
        arquivo_municipios: Caminho para o arquivo JSON de municípios
        
    Returns:
        Dicionário com código do município como chave (int) e dicionário com 'nome' e 'sigla_uf' como valor
    """
    caminho_municipios = Path(arquivo_municipios)
    
    try:
        with open(caminho_municipios, 'r', encoding='utf-8') as f:
            municipios = json.load(f)
        
        # Cria um dicionário indexado por código (convertido para int)
        # Os códigos no JSON podem vir como string, então convertemos para int
        municipios_dict = {}
        for municipio in municipios:
            codigo = municipio['codigo']
            # Converte para int se for string
            codigo_int = int(codigo) if isinstance(codigo, str) else codigo
            nome = municipio['nome']
            sigla_uf = municipio.get('sigla_uf', '')
            # Armazena tanto o nome quanto a sigla
            municipios_dict[codigo_int] = {
                'nome': nome,
                'sigla_uf': sigla_uf
            }
        
        return municipios_dict
    except FileNotFoundError:
        print(f"Arquivo '{arquivo_municipios}' não encontrado. Municípios não serão resolvidos.")
        return {}
    except Exception as e:
        print(f"Erro ao carregar municípios: {e}")
        return {}


def obter_nome_cidade(codigo_municipio, municipios_dict):
    """
    Obtém apenas o nome da cidade baseado no código.
    
    Args:
        codigo_municipio: Código do município (int ou string)
        municipios_dict: Dicionário de municípios carregado
        
    Returns:
        Nome da cidade ou "codigo - Municipio não identificado" se não encontrado
    """
    if codigo_municipio is None:
        return ""
    
    # Converte para int para buscar no dicionário
    try:
        codigo_int = int(codigo_municipio) if not isinstance(codigo_municipio, int) else codigo_municipio
        municipio_info = municipios_dict.get(codigo_int)
        
        if municipio_info:
            nome = municipio_info.get('nome', '')
            return nome if nome else f"{codigo_municipio} - Municipio não identificado"
        else:
            return f"{codigo_municipio} - Municipio não identificado"
    except (ValueError, TypeError):
        # Se não conseguir converter, retorna mensagem de erro
        return f"{codigo_municipio} - Municipio não identificado"


def obter_sigla_uf(codigo_municipio, municipios_dict):
    """
    Obtém apenas a sigla da UF baseado no código do município.
    
    Args:
        codigo_municipio: Código do município (int ou string)
        municipios_dict: Dicionário de municípios carregado
        
    Returns:
        Sigla da UF ou string vazia se não encontrado
    """
    if codigo_municipio is None:
        return ""
    
    # Converte para int para buscar no dicionário
    try:
        codigo_int = int(codigo_municipio) if not isinstance(codigo_municipio, int) else codigo_municipio
        municipio_info = municipios_dict.get(codigo_int)
        
        if municipio_info:
            sigla_uf = municipio_info.get('sigla_uf', '')
            return sigla_uf
        else:
            return ""
    except (ValueError, TypeError):
        # Se não conseguir converter, retorna string vazia
        return ""


def converter_utc_para_sp(data_utc_str):
    """
    Converte uma data/hora em UTC para o fuso horário de São Paulo
    e formata no padrão brasileiro.
    
    Args:
        data_utc_str: String com data/hora em UTC (formato ISO 8601, ex: "2026-03-01T02:20:31Z")
        
    Returns:
        String formatada no padrão brasileiro (DD/MM/YYYY HH:MM:SS) ou None se não conseguir converter
    """
    if not data_utc_str:
        return None
    
    try:
        # Normaliza a string removendo o 'Z' e adicionando '+00:00' para indicar UTC
        if data_utc_str.endswith('Z'):
            data_utc_str = data_utc_str[:-1] + '+00:00'
        elif '+' not in data_utc_str and data_utc_str.count('-') >= 3:
            # Se não tem timezone, assume UTC
            data_utc_str = data_utc_str + '+00:00'
        
        # Usa fromisoformat que é mais robusto
        dt_utc = datetime.fromisoformat(data_utc_str)
        
        # Se não tem timezone, assume UTC
        if dt_utc.tzinfo is None:
            try:
                # Tenta usar zoneinfo (Python 3.9+)
                from zoneinfo import ZoneInfo
                dt_utc = dt_utc.replace(tzinfo=ZoneInfo('UTC'))
            except ImportError:
                # Fallback para pytz
                import pytz
                dt_utc = pytz.UTC.localize(dt_utc)
        
        # Converte para fuso horário de São Paulo
        try:
            # Tenta usar zoneinfo primeiro (Python 3.9+)
            from zoneinfo import ZoneInfo
            tz_sp = ZoneInfo('America/Sao_Paulo')
            dt_sp = dt_utc.astimezone(tz_sp)
        except (ImportError, TypeError, AttributeError):
            # Fallback para pytz
            import pytz
            tz_sp = pytz.timezone('America/Sao_Paulo')
            dt_sp = dt_utc.astimezone(tz_sp)
        
        # Formata no padrão brasileiro: DD/MM/YYYY HH:MM:SS
        return dt_sp.strftime('%d/%m/%Y %H:%M:%S')
        
    except Exception as e:
        # Se houver erro na conversão, retorna None (silenciosamente)
        # Descomente a linha abaixo para debug se necessário
        # print(f"Erro ao converter data '{data_utc_str}': {e}")
        return None


def exportar_para_xlsx(arquivo_json, arquivo_xlsx, nomes_colunas, arquivo_municipios='municipios.json'):
    """
    Exporta os registros do JSON para um arquivo XLSX.
    
    Args:
        arquivo_json: Caminho para o arquivo JSON de entrada
        arquivo_xlsx: Caminho para o arquivo XLSX de saída
        nomes_colunas: Dicionário mapeando campos do JSON para nomes das colunas
        arquivo_municipios: Caminho para o arquivo JSON de municípios
    """
    caminho_json = Path(arquivo_json)
    
    # Carrega o JSON
    with open(caminho_json, 'r', encoding='utf-8') as f:
        dados_json = json.load(f)
    
    # Carrega os municípios uma vez para uso em todas as linhas
    municipios_dict = carregar_municipios(arquivo_municipios)
    
    # Cria o workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Débitos Processados"
    
    # Define a ordem das colunas baseada no dicionário nomes_colunas
    campos_ordenados = list(nomes_colunas.keys())
    titulos_colunas = [nomes_colunas[campo] for campo in campos_ordenados]
    
    # Escreve o cabeçalho
    for col_idx, titulo in enumerate(titulos_colunas, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = titulo
        cell.font = Font(bold=True)
    
    # Função auxiliar para obter valor de um campo (pode estar no nível raiz ou dentro de 'cte')
    def obter_valor(registro, campo):
        # Primeiro tenta no nível raiz
        if campo in registro:
            return registro[campo]
        
        # Depois tenta dentro de 'cte' se existir
        if 'cte' in registro and isinstance(registro['cte'], dict):
            if campo in registro['cte']:
                return registro['cte'][campo]
        
        # Tratamento especial para campos com nomes diferentes
        if campo == 'ValorNaoPago':
            # Tenta 'valorNaoPago' no nível raiz
            if 'valorNaoPago' in registro:
                return registro['valorNaoPago']
            if 'cte' in registro and isinstance(registro['cte'], dict):
                if 'valorNaoPago' in registro['cte']:
                    return registro['cte']['valorNaoPago']
        
        return None
    
    # Escreve os dados
    for row_idx, registro in enumerate(dados_json['dados'], start=2):
        for col_idx, campo in enumerate(campos_ordenados, start=1):
            # Se for campo de UF, busca diretamente pelo código do município correspondente
            if campo == 'uf_orig':
                codigo_orig = obter_valor(registro, 'cd_mun_orig')
                sigla_uf = obter_sigla_uf(codigo_orig, municipios_dict)
                ws.cell(row=row_idx, column=col_idx).value = sigla_uf
            elif campo == 'uf_dest':
                codigo_dest = obter_valor(registro, 'cd_mun_dest')
                sigla_uf = obter_sigla_uf(codigo_dest, municipios_dict)
                ws.cell(row=row_idx, column=col_idx).value = sigla_uf
            else:
                valor = obter_valor(registro, campo)
                
                # Trata valores None ou vazios
                if valor is None:
                    ws.cell(row=row_idx, column=col_idx).value = ""
                else:
                    # Se for o campo dataEmissao, converte de UTC para SP antes de escrever
                    if campo == 'dataEmissao':
                        valor_convertido = converter_utc_para_sp(valor)
                        ws.cell(row=row_idx, column=col_idx).value = valor_convertido if valor_convertido else ""
                    # Se for cd_mun_orig ou cd_mun_dest, busca apenas o nome da cidade
                    elif campo in ('cd_mun_orig', 'cd_mun_dest'):
                        nome_cidade = obter_nome_cidade(valor, municipios_dict)
                        ws.cell(row=row_idx, column=col_idx).value = nome_cidade
                    else:
                        ws.cell(row=row_idx, column=col_idx).value = valor
    
    # Ajusta a largura das colunas
    for col_idx, titulo in enumerate(titulos_colunas, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Define largura mínima baseada no título ou 15 caracteres
        largura = max(len(str(titulo)), 15)
        ws.column_dimensions[col_letter].width = largura
    
    # Salva o arquivo
    caminho_xlsx = Path(arquivo_xlsx)
    wb.save(caminho_xlsx)
    
    print(f"Arquivo XLSX '{arquivo_xlsx}' criado com sucesso!")
    print(f"Total de registros exportados: {len(dados_json['dados'])}")


if __name__ == "__main__":
    crud = CteCRUD()

    arquivo = 'debitos_processados.json'
    
    chaves = extrair_id_documentos_fiscais(arquivo)
    
    ctess_encontrados = crud.get_ids_by_chaves_cte(chaves)
    print(f"Total de CTEs encontrados: {len(ctess_encontrados)}")
    
    caminho_json = Path(arquivo)
    with open(caminho_json, 'r', encoding='utf-8') as f:
        dados_json = json.load(f)
    
    ctes_por_chave = {cte['nr_chave_cte']: cte for cte in ctess_encontrados}
    
    registros_atualizados = 0
    for registro in dados_json['dados']:
        id_documento_fiscal = registro['idDocumentoFiscal']
        
        if id_documento_fiscal in ctes_por_chave:
            cte_encontrado = ctes_por_chave[id_documento_fiscal]
            
            registro['cte'] = {
                'cd_mun_orig': cte_encontrado['cd_mun_orig'],
                'cd_mun_dest': cte_encontrado['cd_mun_dest'],
                'vl_total': cte_encontrado['vl_total'],
                'nr_cnpj_cpf_pagador': f"{cte_encontrado['nr_cnpj_cpf_pagador']} - {cte_encontrado['nm_pessoa_pagador']}",
                'in_exportacao': cte_encontrado['in_exportacao'],
                'tp_destino': cte_encontrado['tp_destino'],
                'nr_chave_nfe': cte_encontrado['nr_chave_nfe'],
                'tipoDocumento': "CT-e",
                'numeroDocumento': cte_encontrado['nr_doc'],
                'numeroDocumentoReferenciado': cte_encontrado['nr_chave_cte_anterior'],
                # 'seguro': "0.00",
                # 'frete': "0.00"
            }
            registros_atualizados += 1
    
    # Salva o JSON atualizado
    with open(caminho_json, 'w', encoding='utf-8') as f:
        json.dump(dados_json, f, ensure_ascii=False, indent=4)
    
    print(f"Total de registros atualizados: {registros_atualizados}")
    print(f"Arquivo '{arquivo}' atualizado com sucesso!")
    
    arquivo_xlsx = 'debitos_processados.xlsx'
    exportar_para_xlsx(arquivo, arquivo_xlsx, nomes_campos_colunas)


